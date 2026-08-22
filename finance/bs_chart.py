#!/usr/bin/env python3
"""
资产负债表 → 16 项分组柱状图 (HTML + PNG)

通用实现：
  - 从任意 BS Excel 读数（自动识别列、字段、单位）
  - 应用用户定义的 16 项合并规则
  - 输出：交互 HTML（Chart.js, 深色主题）+ PNG（matplotlib）

依赖：
  pip install openpyxl matplotlib
"""

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore[assignment]

try:
    import xlrd
except ImportError:
    xlrd = None  # type: ignore[assignment]

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib import font_manager
    HAS_MPL = True
except ImportError:
    HAS_MPL = False

# ──────────────────────────────────────────────────────────────────
#  1. 16 项合并规则（用户预设，可被 --rules 覆盖）
# ──────────────────────────────────────────────────────────────────

DEFAULT_RULES = {
    # 蓝色 — 资产侧（9 项）
    "总现金":      ["货币资金", "结算备付金", "交易性金融资产", "其他应收款合计", "其他应收款"],
    "应收款":      ["应收票据及应收账款", "应收票据及应收账款(元)", "应收款项融资", "应收票据", "应收账款"],
    "预付款":      ["预付款项", "预付账款"],
    "存货":        ["存货"],
    "其他流动":    ["其他流动资产", "一年内到期的非流动资产", "买入返售金融资产"],
    "长期投资":    ["债权投资", "长期应收款", "长期股权投资", "其他权益工具投资"],
    "固定资产":    ["固定资产", "固定资产合计", "在建工程", "在建工程合计"],
    "无形资产":    ["无形资产", "生物性资产", "使用权资产", "商誉", "开发支出"],
    "其他固定":    ["长期待摊费用", "递延所得税资产", "其他非流动资产",
                  "投资性房地产", "其他债权投资", "其他非流动金融资产"],

    # 红色 — 负债侧（7 项）
    "短期借款":    ["短期借款", "一年内到期的非流动负债"],
    "应付款":      ["应付票据及应付账款", "应付票据及应付账款(元)", "其他应付款合计", "其他应付款",
                  "应付票据", "应付账款"],
    "预收款":      ["预收款项", "合同负债", "预收账款"],
    "薪酬&税":     ["应付职工薪酬", "应交税费"],
    "其他负债":    ["其他流动负债", "应付手续费及佣金", "代理买卖证券款", "代理承销证券款"],
    "长期借款":    ["长期借款", "应付债券", "租赁负债"],
    "其他长期":    ["长期应付款合计", "长期应付款", "递延所得税负债",
                  "递延收益-非流动负债", "递延收益", "预计负债", "其他非流动负债"],
}

# 字段名清洗：去掉"(元)"、"(亿元)"、全角括号、空格
FIELD_CLEAN = re.compile(r"[（(].*?[)）]|\s+")

def norm(s: str) -> str:
    return FIELD_CLEAN.sub("", str(s or "")).strip()

def parse_num(v: Any) -> float:
    """统一数值解析：None / '--' / '' 视为 0；带千分位和单位的字符串转 float。"""
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace(",", "")
    if not s or s in ("--", "-", "元", "亿元", "N/A"): return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0

# ──────────────────────────────────────────────────────────────────
#  2. Excel 读取（找行项目列 + 数值列）
# ──────────────────────────────────────────────────────────────────

def _is_date_header(s: str) -> bool:
    """判断表头单元格是否为日期格式（如 2026-03-31 / 2026年12月31日 / 2026/3/31）"""
    if not s:
        return False
    t = str(s)
    return bool(re.search(r"\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}", t)) or \
           bool(re.search(r"\d{4}[-/.年]\d{1,2}", t))


def _parse_period(s) -> str:
    """把表头日期统一成 YYYY-MM-DD 字符串，用于排序和展示。"""
    if not s:
        return ""
    t = str(s).strip()
    m = re.match(r"(\d{4})[-/.年](\d{1,2})(?:[-/.月](\d{1,2}))?", t)
    if not m:
        return t
    y, mo, d = m.group(1), int(m.group(2)), m.group(3)
    if d is None:
        # 仅有年月时按月末补全
        last_day = {1:31,2:28,3:31,4:30,5:31,6:30,7:31,8:31,9:30,10:31,11:30,12:31}
        d = last_day.get(mo, 30)
        if mo == 2 and int(y) % 4 == 0 and (int(y) % 100 != 0 or int(y) % 400 == 0):
            d = 29
    return f"{y}-{mo:02d}-{int(d):02d}"


def read_bs(path: Path, unit_hint: str = "auto") -> dict:
    """
    读取资产负债表 Excel，自动选取最新报告期。

    支持两种文件格式：
      - .xls  → xlrd（含 ignore_workbook_corruption 兼容 OLE2 误报）
      - .xlsx → openpyxl

    多报告期文件（表头含多列日期）时，按日期排序取最新一列作为数值列。
    单列文件（仅一列日期或仅一列数值）时，取该列。

    返回：
      { "报告期": "YYYY-MM-DD", "rows": [(原始字段名, 清洗后字段名, 数值)] }
    数值单位已统一为「元」。字段名用清洗后形式作为 key。
    """
    suffix = path.suffix.lower()
    if suffix == ".xls":
        if xlrd is None:
            sys.exit("缺少依赖：pip install xlrd")
        wb = xlrd.open_workbook(str(path), formatting_info=False,
                                ignore_workbook_corruption=True)
        ws = wb.sheet_by_index(0)
        nrows, ncols = ws.nrows, ws.ncols
        def cell_value(r: int, c: int):
            if r >= nrows or c >= ncols:
                return None
            return ws.cell_value(r, c)
        data_rows = nrows
    else:
        if openpyxl is None:
            sys.exit("缺少依赖：pip install openpyxl")
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        _all = list(ws.iter_rows(min_row=1, values_only=True))
        nrows = len(_all)
        ncols = ws.max_column or (len(_all[0]) if _all else 0)
        def cell_value(r: int, c: int):
            if r >= nrows or c >= ncols:
                return None
            return _all[r][c]
        data_rows = nrows

    # 自动探测表头行：跳过开头全空行，找到第一个「首列非空 且 含至少一个日期表头」的行
    header_row = 0
    for r in range(min(5, data_rows)):
        first_cell = cell_value(r, 0)
        if first_cell is None or str(first_cell).strip() == "":
            continue
        # 检查该行是否有日期型表头
        has_date = any(
            _is_date_header(str(cell_value(r, j) or ""))
            for j in range(1, ncols)
        )
        if has_date:
            header_row = r
            break
    else:
        # 兜底：找第一个非空行作为表头
        for r in range(min(5, data_rows)):
            if cell_value(r, 0) not in (None, ""):
                header_row = r
                break

    header = [cell_value(header_row, j) for j in range(ncols)]
    data_start = header_row + 1

    # 科目列：默认第 0 列
    label_col = 0

    # 找所有「日期型」表头列，按日期排序取最新
    date_cols = []  # [(parsed_period, col_index, raw_header)]
    for j, h in enumerate(header):
        if j == label_col:
            continue
        if h and _is_date_header(str(h)):
            date_cols.append((_parse_period(h), j, str(h)))

    if date_cols:
        # 按解析后的日期字符串降序，取最新
        date_cols.sort(key=lambda x: x[0], reverse=True)
        value_col = date_cols[0][1]
        period_str = date_cols[0][0]
    else:
        # 兜底：没有日期表头时，取第一个非空、非科目列的列
        value_col = 1 if ncols > 1 else 0
        period_str = str(header[value_col]) if value_col < ncols else ""

    # 单位推断
    if unit_hint == "auto":
        unit_hint = "元"
    unit_div = {"元": 1.0, "万元": 1e4, "亿元": 1e8}.get(unit_hint, 1.0)

    rows = []
    SKIP_PREFIX = ("其中", "其中：", "减:", "减：", "加:", "加：")
    for r in range(data_start, data_rows):
        raw_label = cell_value(r, label_col)
        if raw_label is None or str(raw_label).strip() == "":
            continue
        raw_label = str(raw_label).strip()
        # 跳过"其中..."(子明细)与"减..."等附注行 — 这些会与合计行重复计入
        if any(raw_label.startswith(p) for p in SKIP_PREFIX):
            continue
        val = cell_value(r, value_col)
        if val is None or str(val).strip() == "":
            continue
        v = parse_num(val)
        if v == 0:
            continue
        rows.append((raw_label, norm(raw_label), v * unit_div))

    return {"报告期": period_str, "rows": rows}


# ──────────────────────────────────────────────────────────────────
#  3. 应用合并规则
# ──────────────────────────────────────────────────────────────────

def apply_rules(bs: dict, rules: dict, prefer_precomputed: bool = True) -> dict:
    """
    按 rules 把原始行项目合并为 16 项。

    求和优先：先对源科目求和（"及"类合并科目替代其自身组成部分）；
    仅当求和为 0（源科目全部缺失/为零）时，才兜底使用 Excel 中与分组名
    精确同名的「预计算总计行」（字段名 = 分组名 + 可选单位后缀）。
    """
    # 索引：清洗后的字段名 → 数值
    raw_index = {lab_n: v for _, lab_n, v in bs["rows"]}
    # 也建一个 raw → norm 的索引（兜底）
    raw_index_by_full = {lab_raw: v for lab_raw, _, v in bs["rows"]}

    def find_val(candidates: list[str]) -> float:
        """求值策略：
        1) 含"及"的合并科目（如 应收票据及应收账款）优先 — 用它替代自身的
           组成部分（按名称拆"及"），避免与子项双重计算；
        2) 其余未被覆盖的源字段独立累加（按 norm 名去重防重复加）
        """
        seen = set()
        uniq = []
        for c in candidates:
            cn = norm(c)
            if cn in seen: continue
            seen.add(cn)
            uniq.append(cn)

        total = 0.0
        covered: set[str] = set()   # 已被"及"合并科目覆盖的组成部分名
        for cn in uniq:
            if cn in raw_index and "及" in cn:
                total += raw_index[cn]
                covered.update(p for p in cn.split("及") if p)
        for cn in uniq:
            if cn in covered or (cn in raw_index and "及" in cn):
                continue
            total += raw_index.get(cn, 0)
        return total

    result = {}
    for grp_name, sources in rules.items():
        merged = find_val(sources)
        if merged == 0 and prefer_precomputed:
            # 兜底：源科目全部缺失/为零时，才使用表内与分组名同名的预计算行
            merged = find_val([grp_name, grp_name + "(元)", grp_name + "(亿元)"])
        result[grp_name] = merged
    return result


# ──────────────────────────────────────────────────────────────────
#  4. 出图：HTML (Chart.js) + PNG (matplotlib)
# ──────────────────────────────────────────────────────────────────

LABELS = list(DEFAULT_RULES.keys())  # 9 asset + 7 liab（保持用户给定顺序）

ASSET_LABELS = LABELS[:9]
LIAB_LABELS = LABELS[9:]

def render_html(values: dict, period: str, asset_color="#3a7ecf",
                liab_color="#d63031", title="资产负债结构") -> str:
    """生成深色主题的 Chart.js HTML（含 datalabels，可在浏览器打开）"""
    vals_yi = {k: round(v / 1e8, 4) for k, v in values.items()}
    max_val = max(vals_yi.values()) if vals_yi else 1
    ymax = ((max_val // 5) + 1) * 5 if max_val > 0 else 5
    ymax = max(ymax, 5)

    labels_js = json.dumps(LABELS, ensure_ascii=False)
    data_js = json.dumps([vals_yi.get(k, 0) for k in LABELS], ensure_ascii=False)
    title_js = json.dumps(f"{title} ({period})", ensure_ascii=False)

    return f"""<!doctype html><html lang="zh"><head><meta charset="utf-8">
<title>{title} ({period})</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/chartjs-plugin-datalabels/2.2.0/chartjs-plugin-datalabels.min.js"></script>
<style>
  body{{margin:0;background:#1e1e1e;color:#ddd;font-family:-apple-system,BlinkMacSystemFont,'Microsoft YaHei',sans-serif}}
  .wrap{{max-width:1200px;margin:24px auto;padding:0 24px}}
  h1{{font-size:18px;font-weight:500;text-align:center;margin:0 0 16px;color:#fff}}
  .legend{{display:flex;gap:24px;justify-content:center;margin-bottom:12px;font-size:13px}}
  .legend i{{display:inline-block;width:12px;height:12px;border-radius:2px;margin-right:6px;vertical-align:middle}}
  .asset i{{background:{asset_color}}} .liab i{{background:{liab_color}}}
  .card{{background:#2a2a2a;border-radius:8px;padding:16px}}
</style></head><body>
<div class="wrap">
  <h1 id="t"></h1>
  <div class="legend">
    <span class="asset"><i></i>资产侧</span>
    <span class="liab"><i></i>负债侧</span>
  </div>
  <div class="card"><canvas id="c" style="max-height:520px"></canvas></div>
  <p style="text-align:center;color:#888;font-size:12px;margin-top:12px">
    单位：亿元 ｜ 数据来源：用户上传 Excel ｜ 16 项按预设规则合并
  </p>
</div>
<script>
document.getElementById('t').textContent = {title_js};
const labels = {labels_js};
const vals = {data_js};
Chart.register(ChartDataLabels);
new Chart(document.getElementById('c'), {{
  type:'bar', data:{{ labels, datasets:[{{
    label:'金额(亿元)', data:vals,
    backgroundColor:(ctx)=>ctx.dataIndex<9?'{asset_color}':'{liab_color}',
    borderWidth:0
  }}]}},
  options:{{
    responsive:true, maintainAspectRatio:false,
    layout:{{padding:{{top:28}}}},
    plugins:{{
      legend:{{display:false}},
      tooltip:{{callbacks:{{label:c=>c.parsed.y.toFixed(3)+' 亿元'}}}},
      datalabels:{{
        anchor:'end', align:'top', offset:2,
        color:'#fff', font:{{size:11,weight:500}},
        formatter:v=>v===0?'':v.toFixed(2),
        display:c=>vals[c.dataIndex]>0
      }}
    }},
    scales:{{
      x:{{
        ticks:{{color:'#ccc',font:{{size:11}},autoSkip:false,maxRotation:45,minRotation:45}},
        grid:{{display:false}}
      }},
      y:{{
        beginAtZero:true, max:{ymax},
        ticks:{{color:'#ccc',font:{{size:11}}}},
        grid:{{color:'rgba(255,255,255,0.08)'}}
      }}
    }}
  }}
}});
</script></body></html>
"""


def render_png(values: dict, period: str, out: Path,
               title="资产负债结构") -> None:
    """matplotlib 静态图（深色主题）。用于嵌 PPT 或报告。"""
    if not HAS_MPL:
        print("⚠ matplotlib 未安装，跳过 PNG 输出", file=sys.stderr)
        return
    # 中文字体：优先 Microsoft YaHei / SimHei，找不到时退到 Noto Sans CJK
    try:
        candidates = ["Microsoft YaHei", "SimHei", "Microsoft JhengHei",
                      "PingFang SC", "Heiti SC", "Noto Sans CJK SC",
                      "Source Han Sans SC", "DejaVu Sans"]
        available = {f.name for f in font_manager.fontManager.ttflist}
        chosen = next((c for c in candidates if c in available), "DejaVu Sans")
        plt.rcParams["font.family"] = [chosen]
        plt.rcParams["font.sans-serif"] = [chosen]
        plt.rcParams["axes.unicode_minus"] = False
    except Exception:
        pass

    vals_yi = [round(values.get(k, 0) / 1e8, 4) for k in LABELS]
    fig, ax = plt.subplots(figsize=(13, 6), facecolor="#1e1e1e")
    ax.set_facecolor("#1e1e1e")
    colors = ["#3a7ecf"] * 9 + ["#d63031"] * 7
    bars = ax.bar(LABELS, vals_yi, color=colors)
    for b, v in zip(bars, vals_yi):
        if v > 0:
            ax.text(b.get_x() + b.get_width()/2, v, f"{v:.2f}",
                    ha="center", va="bottom", color="white", fontsize=10)
    ax.set_title(f"{title} ({period})", color="white", fontsize=16, pad=16)
    ax.set_ylabel("亿元", color="#ccc", fontsize=11)
    ax.tick_params(axis="x", rotation=45, colors="#ccc", labelsize=10)
    ax.tick_params(axis="y", colors="#ccc", labelsize=10)
    for s in ax.spines.values(): s.set_color("#444")
    ax.grid(axis="y", color="#333", linestyle="--", alpha=0.5)
    fig.tight_layout()
    fig.savefig(out, dpi=150, facecolor="#1e1e1e")
    plt.close(fig)


# ──────────────────────────────────────────────────────────────────
#  5. CLI
# ──────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(description="资产负债表 → 16 项分组柱状图")
    p.add_argument("input", help="Excel 文件路径（.xls / .xlsx）")
    p.add_argument("-o", "--output-dir", default=".", help="输出目录（默认当前目录）")
    p.add_argument("--period", default=None, help="覆盖报告期显示文本")
    p.add_argument("--title", default="资产负债结构", help="图表标题")
    p.add_argument("--rules", default=None, help="自定义 JSON 合并规则文件")
    p.add_argument("--no-png", action="store_true", help="不输出 PNG")
    p.add_argument("--no-html", action="store_true", help="不输出 HTML")
    p.add_argument("--unit", default="元", choices=["元", "万元", "亿元"],
                   help="Excel 中数值的单位（默认：元）")
    args = p.parse_args()

    src = Path(args.input)
    if not src.exists(): sys.exit(f"找不到文件：{src}")
    out_dir = Path(args.output_dir); out_dir.mkdir(parents=True, exist_ok=True)

    rules = DEFAULT_RULES
    if args.rules:
        rules = json.loads(Path(args.rules).read_text(encoding="utf-8"))

    bs = read_bs(src, unit_hint=args.unit)
    values = apply_rules(bs, rules)
    period = args.period or bs.get("报告期") or src.stem

    stem = src.stem
    html_path = out_dir / f"{stem}_资产负债结构.html"
    png_path = out_dir / f"{stem}_资产负债结构.png"
    json_path = out_dir / f"{stem}_分组数据.json"

    if not args.no_html:
        html_path.write_text(render_html(values, period, title=args.title),
                             encoding="utf-8")
        print(f"✓ HTML → {html_path}")
    if not args.no_png:
        render_png(values, period, png_path, title=args.title)
        if png_path.exists(): print(f"✓ PNG  → {png_path}")
    json_path.write_text(json.dumps(
        {"期间": period, "分组(亿元)": {k: round(v/1e8, 4) for k, v in values.items()},
         "原始字段数": len(bs["rows"])}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"✓ JSON → {json_path}")


if __name__ == "__main__":
    main()
