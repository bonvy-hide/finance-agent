"""个股财报数据标准化算法模块（纯业务逻辑，无 Web 依赖）。

从原始 .xls（diy_report 模板）读取财报科目，按规则计算滚动 TTM、
归母净资产、自由现金流 FCF，并输出标准化数据结构。

支持两种调用方式：
    1. CLI: `python -m finance.stock_normalize --input xxx.xls --output xxx.xlsx`
       或经 scripts/convert_stock_data.py 薄包装调用
    2. 模块: `from finance.stock_normalize import normalize`
       返回 NormalizedData 字典，供 Web 路由或图表算法使用

设计原则：
    - 不 import 任何 FastAPI / HTTP 模块（四层架构约定）
    - 科目匹配基于名称（顺序可变），不依赖行号
    - 数值底层保留 float（元），单位转换在图表层处理
"""
from __future__ import annotations

import argparse
import sys
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, TypedDict

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Windows 控制台中文/特殊字符兼容
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except (AttributeError, OSError):
        pass


# ---------------------------------------------------------------------------
# 常量定义
# ---------------------------------------------------------------------------

# 输出列固定顺序（行=时间，列=指标）
# 顺序：时间列 → 4 个滚动 TTM 列 → 归母净资产 → 5 个单期金额列 → 自由现金流FCF
#       → 6 个新增科目（营业成本/销售毛利率/销售费用/管理费用/研发费用/应付票据及应付账款）
OUTPUT_COLUMNS: List[str] = [
    "科目\\时间",
    "滚动总营收",
    "滚动现金流",
    "滚动净利润",
    "滚动现金流净额",
    "归母净资产",
    "总营收",
    "主业现金流",
    "净利润",
    "现金流净额",
    "自由现金流FCF",
    # 新增 6 列（单期值，不算 TTM）
    "营业成本",
    "销售毛利率",
    "销售费用",
    "管理费用",
    "研发费用",
    "应付票据及应付账款",
]

# 原始科目名 → 内部别名映射（原有 7 项 + 新增 6 项）
LABEL_MAP: Dict[str, str] = {
    # 原有科目
    "总营收": "营业总收入(元)",
    "净利润": "净利润(元)",
    "主业现金流": "销售商品、提供劳务收到的现金(元)",
    "现金流净额": "经营活动产生的现金流量净额(元)",
    "每股净资产": "每股净资产(元)",
    "实收资本": "实收资本（或股本）(元)",
    "资本支出": "购建固定资产、无形资产和其他长期资产支付的现金(元)",
    # 新增科目（原始科目名 → 内部别名）
    "营业成本": "其中：营业成本(元)",
    "销售毛利率": "销售毛利率",
    "销售费用": "销售费用(元)",
    "管理费用": "管理费用(元)",
    "研发费用": "研发费用(元)",
    "应付票据及应付账款": "应付票据及应付账款(元)",
}

# 需要计算 TTM 的科目（内部别名）——仅原有 4 个金额科目，新增科目不算 TTM
TTM_LABELS: List[str] = ["总营收", "主业现金流", "净利润", "现金流净额"]

# TTM 输出列名 → 内部别名映射
TTM_COLUMN_MAP: Dict[str, str] = {
    "滚动总营收": "总营收",
    "滚动现金流": "主业现金流",
    "滚动净利润": "净利润",
    "滚动现金流净额": "现金流净额",
}

# 新增科目（单期值）的输出列名 → 内部别名映射
SINGLE_PERIOD_NEW_COLUMNS: Dict[str, str] = {
    "营业成本": "营业成本",
    "销售毛利率": "销售毛利率",
    "销售费用": "销售费用",
    "管理费用": "管理费用",
    "研发费用": "研发费用",
    "应付票据及应付账款": "应付票据及应付账款",
}

# 毛利率类科目（百分比，需要转小数）
PERCENT_LABELS: List[str] = ["销售毛利率"]

# 数据范围
BUFFER_START_YEAR = 2015  # 缓冲起始年（含 2015 全年用于 TTM 计算）
OUTPUT_START_YEAR = 2016  # 最终输出起始年（10 年完整周期）

# 项目根目录（本模块位于 <root>/finance/ 下）
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = PROJECT_ROOT / "static" / "template" / "002922_diy_report (6).xls"
DEFAULT_OUTPUT = PROJECT_ROOT / "outputs" / "主要数据_标准化.xlsx"


# ---------------------------------------------------------------------------
# 类型定义（供 Web 层类型提示用，运行时为 dict）
# ---------------------------------------------------------------------------


class NormalizedData(TypedDict):
    """标准化数据的统一结构，供 Web 路由和图表算法消费。

    Attributes:
        periods: 报告期字符串列表 (YYYY-MM-DD)，升序
        columns: 列名列表，与 OUTPUT_COLUMNS 一致（首列为 "科目\\时间"）
        rows: 数据行，每行对应一个报告期，长度 = len(columns)
              首列为日期字符串，其余为 float 或 None
        meta: 元信息（科目数、报告期数、TTM 列、新增列等）
    """

    periods: List[str]
    columns: List[str]
    rows: List[List[Optional[float]]]
    meta: Dict[str, Any]


# ---------------------------------------------------------------------------
# 1. 读取 xls
# ---------------------------------------------------------------------------


def _to_float(value: Any) -> Optional[float]:
    """把 xlrd 单元格值转为 float，空值/非数字/缺失值返回 None。

    处理以下情况：
        - None / 空字符串 → None
        - int / float → float(value)
        - 字符串数字（如 "1234.56"）→ float
        - 字符串 '--'（缺失标记）→ None
        - 百分比字符串（如 "19.82%"）→ 交给 _parse_percent 处理（这里不处理）
        - 其他无法解析的字符串 → None
    """
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        s = str(value).strip()
        if s == "" or s == "--":
            return None
        return float(s)
    except (ValueError, TypeError):
        return None


def _parse_percent(value: Any) -> Optional[float]:
    """把百分比单元格值转为小数 float。

    xlrd 读取百分比单元格：
        - 若单元格格式为百分比，xlrd 已自动转小数（如 0.1982），直接返回
        - 若以字符串形式存储（如 "19.82%"），strip 后除以 100
        - 其他情况回退到 _to_float
    """
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        # xlrd 已转小数，但保险起见检查范围（百分比应在 [-1, 1] 之外才需除 100）
        # 通常财报毛利率百分比单元格 xlrd 直接返回小数（如 0.1982）
        return float(value)
    s = str(value).strip()
    if s == "" or s == "--":
        return None
    if s.endswith("%"):
        try:
            return float(s[:-1].strip()) / 100.0
        except ValueError:
            return None
    # 非百分比字符串，尝试普通 float
    try:
        return float(s)
    except ValueError:
        return None


def _parse_period(period_str: str) -> Optional[date]:
    """解析 'YYYY-MM-DD' 字符串为 date 对象。"""
    period_str = str(period_str).strip()
    try:
        y, m, d = period_str.split("-")
        return date(int(y), int(m), int(d))
    except (ValueError, AttributeError):
        return None


def _quarter_of(d: date) -> int:
    """根据月份返回季度编号（1-4）。"""
    return (d.month - 1) // 3 + 1


def read_stock_xls(path: Path) -> Tuple[Dict[str, Dict[date, float]], List[date]]:
    """读取个股财报 .xls 文件，返回结构化数据。

    基于科目名匹配（顺序可变），不依赖行号。
    自动跳过"归母净资产"和"自由现金流FCF"行（这两个由 compute_derived 计算，
    模板中不需要手动添加；即使添加了也会被忽略，避免与计算值冲突）。

    Args:
        path: .xls 文件路径

    Returns:
        (data, periods) 二元组：
        - data: {原始科目名: {报告期date: 数值float}}，数值缺失则该 date 不在 dict 中
        - periods: 报告期 date 列表（升序，已去重、去空）
    """
    # 由 compute_derived 计算的科目，模板中不需要也不应该手动添加
    DERIVED_LABELS = {"归母净资产", "自由现金流FCF"}

    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_index(0)

    if ws.nrows < 3 or ws.ncols < 2:
        raise ValueError(f"文件结构异常：nrows={ws.nrows}, ncols={ws.ncols}")

    # 表头在第 1 行（0-indexed），A1='科目\\时间'，B1 起为报告期
    header_row = 1
    label_col = 0

    # 解析报告期表头（第 1 列起）
    periods_set: Dict[date, int] = {}  # date -> 列索引
    for col in range(1, ws.ncols):
        raw = ws.cell_value(header_row, col)
        d = _parse_period(str(raw))
        if d is not None:
            periods_set[d] = col

    if not periods_set:
        raise ValueError("未解析到任何报告期表头")

    periods: List[date] = sorted(periods_set.keys())

    # 读取科目行（第 2 行起），按科目名存入 dict（顺序无关）
    data: Dict[str, Dict[date, float]] = {}
    for row in range(header_row + 1, ws.nrows):
        label = ws.cell_value(row, label_col)
        if label is None or str(label).strip() == "":
            continue
        label = str(label).strip()

        # 跳过计算列（归母净资产、自由现金流FCF）——这些由 compute_derived 计算，
        # 模板中不需要手动添加；即使添加了也忽略，避免与计算值冲突
        if label in DERIVED_LABELS:
            continue

        col_values: Dict[date, float] = {}
        for d, col in periods_set.items():
            raw = ws.cell_value(row, col)
            # 毛利率类科目用百分比解析，其他用普通 float
            if label in PERCENT_LABELS:
                v = _parse_percent(raw)
            else:
                v = _to_float(raw)
            if v is not None:
                col_values[d] = v

        # 即使该行所有期都为空，也保留科目键（值为空 dict），
        # 便于后续 build_output_rows 统一处理缺失
        data[label] = col_values

    return data, periods


# ---------------------------------------------------------------------------
# 2. 计算科目（归母净资产、FCF）
# ---------------------------------------------------------------------------


def compute_derived(data: Dict[str, Dict[date, float]]) -> Dict[str, Dict[date, float]]:
    """计算归母净资产与自由现金流 FCF，写入 data 并返回。

    归母净资产 = 每股净资产 × 实收资本（时点值）
    FCF = 经营活动现金流量净额 - 购建固定资产...支付的现金（单期累计值）
    """
    eps_name = LABEL_MAP["每股净资产"]
    capital_name = LABEL_MAP["实收资本"]
    ocf_name = LABEL_MAP["现金流净额"]
    capex_name = LABEL_MAP["资本支出"]

    eps_map = data.get(eps_name, {})
    capital_map = data.get(capital_name, {})
    ocf_map = data.get(ocf_name, {})
    capex_map = data.get(capex_name, {})

    equity_map: Dict[date, float] = {}
    fcf_map: Dict[date, float] = {}

    all_dates = set(eps_map.keys()) | set(capital_map.keys()) | set(ocf_map.keys()) | set(capex_map.keys())
    for d in all_dates:
        eps = eps_map.get(d)
        cap = capital_map.get(d)
        if eps is not None and cap is not None:
            equity_map[d] = eps * cap

        ocf = ocf_map.get(d)
        capex = capex_map.get(d)
        if ocf is not None and capex is not None:
            fcf_map[d] = ocf - capex

    data["归母净资产"] = equity_map
    data["自由现金流FCF"] = fcf_map
    return data


# ---------------------------------------------------------------------------
# 3. 滚动 TTM 计算
# ---------------------------------------------------------------------------


def _same_period_prev_year(periods: List[date], d: date) -> Optional[date]:
    """返回上年同期的 date（存在则返回，否则 None）。"""
    prev = date(d.year - 1, d.month, d.day)
    return prev if prev in periods else None


def _prev_year_end(periods: List[date], d: date) -> Optional[date]:
    """返回上一自然年的 12-31 date（存在则返回，否则 None）。"""
    prev_end = date(d.year - 1, 12, 31)
    return prev_end if prev_end in periods else None


def compute_ttm(
    values: Dict[date, float], periods: List[date]
) -> Dict[date, float]:
    """对指定科目的累计值序列计算每个报告期的滚动 TTM。

    公式（财报数据为累计值）：
        - 年报（12-31）：TTM = values[d]
        - Q1（03-31）：TTM = values[上年12-31] - values[上年03-31] + values[d]
        - Q2（06-30）：TTM = values[d] + (values[上年12-31] - values[上年06-30])
        - Q3（09-30）：TTM = values[d] + (values[上年12-31] - values[上年09-30])

    若任一所需历史数据缺失，该期 TTM 不在返回结果中（视为空值）。
    """
    periods_set = set(periods)
    ttm: Dict[date, float] = {}

    for d in periods:
        q = _quarter_of(d)
        cur = values.get(d)
        if cur is None:
            continue

        if q == 4:
            # 年报：本身即全年累计 = TTM
            ttm[d] = cur
            continue

        prev_end = _prev_year_end(periods_set, d)
        prev_same = _same_period_prev_year(periods_set, d)
        if prev_end is None or prev_same is None:
            continue

        v_prev_end = values.get(prev_end)
        v_prev_same = values.get(prev_same)
        if v_prev_end is None or v_prev_same is None:
            continue

        # TTM = 本期累计 + 上年全年 - 上年同期累计
        ttm[d] = cur + v_prev_end - v_prev_same

    return ttm


# ---------------------------------------------------------------------------
# 4. 组装输出行
# ---------------------------------------------------------------------------


def build_output_rows(
    data: Dict[str, Dict[date, float]],
    periods: List[date],
    output_start_year: int = OUTPUT_START_YEAR,
) -> Tuple[List[date], List[List]]:
    """组装输出数据行（时间升序，按 OUTPUT_COLUMNS 固定列顺序）。

    缺失的季度报告期会自动补全（数据点为 None），使 x 轴显示完整季度序列，
    例如原始数据缺 2016-06-30 时，输出仍包含该期，所有指标为 None。

    Returns:
        (output_periods, rows) 其中 rows 每行长度 = len(OUTPUT_COLUMNS)
    """
    # 预计算每个 TTM 科目（仍用原始 periods，缺失期不参与 TTM）
    ttm_data: Dict[str, Dict[date, float]] = {}
    for alias in TTM_LABELS:
        raw_name = LABEL_MAP[alias]
        ttm_data[alias] = compute_ttm(data.get(raw_name, {}), periods)

    # 过滤输出周期：>= output_start_year（仅原始数据）
    raw_out_periods = [d for d in periods if d.year >= output_start_year]
    if not raw_out_periods:
        return [], []

    # 补全缺失季度：从 output_start_year-Q1 到原始最大期，生成完整季度序列
    # 季度末日期固定为 03-31 / 06-30 / 09-30 / 12-31
    quarter_end_days = [(3, 31), (6, 30), (9, 30), (12, 31)]
    max_period = max(raw_out_periods)
    out_periods: List[date] = []
    y = output_start_year
    q_idx = 0  # 从 Q1 开始
    while True:
        m, d_day = quarter_end_days[q_idx]
        d = date(y, m, d_day)
        out_periods.append(d)
        if d == max_period:
            break
        q_idx += 1
        if q_idx >= 4:
            q_idx = 0
            y += 1
        # 安全保护：避免无限循环
        if y > max_period.year + 1:
            break

    # 原始数据期集合，用于判断某期是否缺失
    raw_period_set = set(raw_out_periods)

    rows: List[List] = []
    for d in out_periods:
        if d in raw_period_set:
            # 真实数据期：正常取值
            row: List = [d.strftime("%Y-%m-%d")]
            # 滚动列（4 列）
            for col_name in ["滚动总营收", "滚动现金流", "滚动净利润", "滚动现金流净额"]:
                alias = TTM_COLUMN_MAP[col_name]
                row.append(ttm_data[alias].get(d))
            # 归母净资产（时点值，计算列）
            row.append(data.get("归母净资产", {}).get(d))
            # 单期值列（4 列，原有科目）
            for alias in ["总营收", "主业现金流", "净利润", "现金流净额"]:
                row.append(data.get(LABEL_MAP[alias], {}).get(d))
            # 自由现金流FCF（计算列）
            row.append(data.get("自由现金流FCF", {}).get(d))
            # 新增 6 列（单期值，从原始 data 按科目名取）
            for out_col, alias in SINGLE_PERIOD_NEW_COLUMNS.items():
                raw_name = LABEL_MAP[alias]
                row.append(data.get(raw_name, {}).get(d))
        else:
            # 补全的缺失期：所有数据点为 None
            row = [d.strftime("%Y-%m-%d")] + [None] * (len(OUTPUT_COLUMNS) - 1)
        rows.append(row)

    return out_periods, rows


# ---------------------------------------------------------------------------
# 5. 高层接口：normalize() 供 Web 调用
# ---------------------------------------------------------------------------


def normalize(path: Path, output_start_year: int = OUTPUT_START_YEAR) -> NormalizedData:
    """读取 .xls 并返回标准化数据结构（供 Web 路由和图表算法使用）。

    Args:
        path: .xls 文件路径
        output_start_year: 输出起始年（默认 2016）

    Returns:
        NormalizedData 字典，含 periods/columns/rows/meta
    """
    data, periods = read_stock_xls(path)
    compute_derived(data)
    out_periods, rows = build_output_rows(data, periods, output_start_year)

    periods_str = [d.strftime("%Y-%m-%d") for d in out_periods]
    meta: Dict[str, Any] = {
        "period_count": len(out_periods),
        "column_count": len(OUTPUT_COLUMNS),
        "ttm_columns": list(TTM_COLUMN_MAP.keys()),
        "new_columns": list(SINGLE_PERIOD_NEW_COLUMNS.keys()),
        "source_file": str(path.name),
        "raw_period_count": len(periods),
    }

    return NormalizedData(
        periods=periods_str,
        columns=list(OUTPUT_COLUMNS),
        rows=rows,
        meta=meta,
    )


# ---------------------------------------------------------------------------
# 6. 写入 xlsx（CLI 模式用）
# ---------------------------------------------------------------------------


def write_xlsx(
    rows: List[List],
    output_path: Path,
    sheet_name: str = "标准化数据",
) -> None:
    """把数据行写入 xlsx，含表头、列宽、数字格式与样式。

    毛利率列（销售毛利率）用百分比格式 0.00%；-0.00%；-，
    其余数值列用 #,##0.00;-#,##0.00;-。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 表头
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F4E78")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(OUTPUT_COLUMNS)
    for col_idx in range(1, len(OUTPUT_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # 数据行
    body_font = Font(name="Arial", size=10)
    first_col_font = Font(name="Arial", size=10, bold=True)
    body_align = Alignment(horizontal="center", vertical="center")
    num_format = "#,##0.00;-#,##0.00;-"
    percent_format = "0.00%;-0.00%;-"
    # 销售毛利率列在 OUTPUT_COLUMNS 中的索引（1-based）
    percent_col_idx = OUTPUT_COLUMNS.index("销售毛利率") + 1

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if c_idx == 1:
                cell.font = first_col_font
                cell.alignment = body_align
            else:
                cell.font = body_font
                cell.alignment = body_align
                if isinstance(val, (int, float)):
                    # 毛利率列用百分比格式，其余用金额格式
                    if c_idx == percent_col_idx:
                        cell.number_format = percent_format
                    else:
                        cell.number_format = num_format

    # 列宽
    ws.column_dimensions["A"].width = 14
    for c_idx in range(2, len(OUTPUT_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 18

    # 冻结首行首列
    ws.freeze_panes = "B2"

    # 行高
    ws.row_dimensions[1].height = 24

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def write_normalized_to_xlsx(
    normalized: NormalizedData,
    output_path: Path,
    sheet_name: str = "标准化数据",
) -> None:
    """把 normalize() 返回的数据结构写入 xlsx（CLI 便捷接口）。"""
    write_xlsx(normalized["rows"], output_path, sheet_name)


# ---------------------------------------------------------------------------
# CLI 入口（可独立运行，也可被 scripts/convert_stock_data.py 调用）
# ---------------------------------------------------------------------------


def main() -> int:
    parser = argparse.ArgumentParser(description="个股财报数据标准化转换")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="输入 .xls 文件路径")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="输出 .xlsx 文件路径")
    args = parser.parse_args()

    input_path: Path = args.input
    output_path: Path = args.output

    if not input_path.exists():
        print(f"[ERROR] 输入文件不存在：{input_path}")
        return 1

    print(f"[1/3] 读取源文件：{input_path}")
    data, periods = read_stock_xls(input_path)
    print(f"      科目数：{len(data)}，报告期数：{len(periods)}")
    print(f"      时间范围：{periods[0]} ~ {periods[-1]}")

    print("[2/3] 计算归母净资产、自由现金流FCF、滚动TTM")
    compute_derived(data)
    out_periods, rows = build_output_rows(data, periods, OUTPUT_START_YEAR)
    print(f"      输出周期：{out_periods[0]} ~ {out_periods[-1]}，共 {len(out_periods)} 期")

    print(f"[3/3] 写入输出文件：{output_path}")
    write_xlsx(rows, output_path)
    print(f"      完成。列：{len(OUTPUT_COLUMNS)}，行：{len(rows) + 1}（含表头）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
