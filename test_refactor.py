"""重构后的端到端验证脚本（测试后删除）"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

import os
import tempfile
import time
import urllib.request
import json
import subprocess
import signal

import openpyxl

print("=" * 60)
print("重构验证测试")
print("=" * 60)

# 1. 构造测试 Excel
print("\n[1] 构造测试 Excel...")
wb = openpyxl.Workbook()
ws = wb.active
ws["A1"] = "科目"
ws["B1"] = "2024-12-31"
test_data = [
    ("货币资金", 10000000000),
    ("应收账款", 5000000000),
    ("存货", 3000000000),
    ("短期借款", 4000000000),
    ("应付账款", 2000000000),
]
for i, (label, val) in enumerate(test_data, start=2):
    ws[f"A{i}"] = label
    ws[f"B{i}"] = val

tmp_xlsx = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
wb.save(tmp_xlsx.name)
tmp_xlsx.close()
print(f"    测试文件: {tmp_xlsx.name}")

# 2. 启动真实服务
print("\n[2] 启动 FastAPI 服务...")
proc = subprocess.Popen(
    ["uv", "run", "uvicorn", "main:app", "--host", "127.0.0.1", "--port", "8765"],
    cwd=r"c:\data\python\finance-agent",
    stdout=subprocess.DEVNULL,
    stderr=subprocess.DEVNULL,
)
time.sleep(5)

try:
    # 3. 测试首页
    print("\n[3] 测试 GET / ...")
    with urllib.request.urlopen("http://127.0.0.1:8765/", timeout=5) as r:
        html = r.read().decode("utf-8")
        assert "资产负债" in html
        print(f"    [OK] 首页返回 {len(html)} 字节")

    # 4. 测试静态文件
    print("\n[4] 测试 GET /static/app.js ...")
    with urllib.request.urlopen("http://127.0.0.1:8765/static/app.js", timeout=5) as r:
        js = r.read().decode("utf-8")
        assert "/api/bs-chart" in js
        print(f"    [OK] app.js 返回 {len(js)} 字节，包含新 API 路径")

    # 5. 测试 API
    print("\n[5] 测试 POST /api/bs-chart ...")
    import urllib.parse

    # 构造 multipart/form-data
    boundary = "----TestBoundary123456"
    with open(tmp_xlsx.name, "rb") as f:
        file_data = f.read()
    body = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="file"; filename="test.xlsx"\r\n'
        f"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
    ).encode() + file_data + f"\r\n--{boundary}--\r\n".encode()

    req = urllib.request.Request(
        "http://127.0.0.1:8765/api/bs-chart",
        data=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=10) as r:
        result = json.loads(r.read().decode("utf-8"))

    print(f"    period: {result.get('period')}")
    print(f"    title: {result.get('title')}")
    print(f"    chart_type: {result.get('chart_type')}")
    print(f"    labels: {result.get('labels')}")
    print(f"    values: {result.get('values')}")
    print(f"    total: {result.get('total')}")
    print(f"    extra: {result.get('extra')}")

    # 断言
    assert result["title"] == "资产负债结构"
    assert result["chart_type"] == "bar"
    assert result["period"] == "2024-12-31"
    assert len(result["labels"]) == 16
    assert len(result["values"]) == 16
    assert result["extra"]["asset_count"] == 9
    assert result["extra"]["raw_rows_count"] == 5
    # 货币资金 100亿
    idx = result["labels"].index("总现金")
    assert abs(result["values"][idx] - 100.0) < 0.01, f"总现金应为100亿, got {result['values'][idx]}"
    print("\n    [OK] 所有断言通过！")

    print("\n" + "=" * 60)
    print("[OK] 重构验证全部通过！")
    print("=" * 60)

except Exception as e:
    print(f"\n[FAIL] 测试失败: {e}")
    raise
finally:
    proc.terminate()
    proc.wait(timeout=5)
    os.unlink(tmp_xlsx.name)
